from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

def exportar_excel_formatado(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Unificado"

    fill_pergunta = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
    fill_cabecalho = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")  # azul claro
    fonte_cabecalho = Font(bold=True)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if r_idx == 1:
                col_letter = cell.column_letter
                ws.column_dimensions[col_letter].width = 50 if c_idx < 30 else 20
                cell.fill = fill_cabecalho
                cell.font = fonte_cabecalho
            else:
                if isinstance(value, str) and value.strip().lower().startswith(("a)", "b)", "c)", "d)", "e)")):
                    cell.fill = fill_pergunta

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output